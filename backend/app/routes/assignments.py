from flask import Blueprint, request, jsonify, current_app
from app.utils.decorators import audit_logger
from flask_jwt_extended import jwt_required, get_jwt_identity
from app import db
from app.models import User, Assignment, AssignmentGroup, AssignmentGroupMember, Workspace, Group, GroupMember, File
from datetime import datetime
import json
import random
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from app.services.livekit_service import get_livekit_service

assignments_bp = Blueprint('assignments', __name__)

# -------------------------------------------------------------------
# ASSIGNMENT MANAGEMENT
# -------------------------------------------------------------------

@assignments_bp.route('/', methods=['GET'])
@jwt_required()
def get_assignments():
    """List assignments for the current user's workspace"""
    current_user_id = get_jwt_identity()
    user = User.query.get(current_user_id)
    
    if not user.workspace_id:
        return jsonify({"error": "No active workspace"}), 400

    query = Assignment.query.filter_by(workspace_id=user.workspace_id)
    
    # Filter by channel if provided
    channel_id = request.args.get('channel_id')
    if channel_id:
        query = query.filter_by(channel_id=channel_id)

    if user.role == 'student':
        query = query.filter_by(status='published')
        
        # Jurisdictional Lockdown: Student MUST be enrolled in the channel
        from app.models import ChannelMember
        enrolled_channel_ids = [m.channel_id for m in ChannelMember.query.filter_by(user_id=user.id).all()]
        
        # If they filtered by channel, check if they're enrolled
        if channel_id and int(channel_id) not in enrolled_channel_ids:
             return jsonify({"error": "You are not enrolled in this course channel"}), 403
             
        # Only show assignments from enrolled channels or public workspace ones
        from sqlalchemy import or_
        query = query.filter(or_(
            Assignment.channel_id.in_(enrolled_channel_ids),
            Assignment.channel_id == None
        ))
        
        # Further scope by channel if applicable
        from app.models import ChannelMember
        student_channel_ids = [m.channel_id for m in ChannelMember.query.filter_by(user_id=user.id).all()]
        query = query.filter(or_(
            Assignment.channel_id.in_(student_channel_ids),
            Assignment.channel_id == None
        ))
    
    assignments = query.all()
    return jsonify([a.to_dict() for a in assignments])

@assignments_bp.route('/', methods=['POST'])
@jwt_required()
@audit_logger('ASSIGNMENT_CREATE', 'assignment')
def create_assignment():
    """Create a new assignment (Lecturer/Admin only)"""
    current_user_id = get_jwt_identity()
    user = User.query.get(current_user_id)
    
    if not user.workspace_id:
        return jsonify({"error": "No active workspace"}), 400
        
    if user.role == 'student':
        return jsonify({"error": "Unauthorized"}), 403
        
    data = request.get_json()
    
    due_date_str = data.get('due_date')
    if due_date_str and isinstance(due_date_str, str):
        due_date_str = due_date_str.replace('Z', '+00:00')
    
    new_assignment = Assignment(
        title=data.get('title'),
        description=data.get('description'),
        workspace_id=user.workspace_id,
        channel_id=data.get('channel_id'),
        created_by=user.id,
        due_date=datetime.fromisoformat(due_date_str) if due_date_str else None,
        status=data.get('status', 'draft'),
        settings=json.dumps(data.get('settings', {}))
    )
    
    db.session.add(new_assignment)
    db.session.commit()
    
    # Secure Collaboration: Notify course/channel room via Socket.IO
    from app import socketio
    room_id = f"channel_{new_assignment.channel_id}" if new_assignment.channel_id else f"ws_{user.workspace_id}"
    socketio.emit('new_assignment', new_assignment.to_dict(), room=room_id)
    
    return jsonify(new_assignment.to_dict()), 201

@assignments_bp.route('/<int:assignment_id>', methods=['GET'])
@jwt_required()
def get_assignment_detail(assignment_id):
    current_user_id = get_jwt_identity()
    user = User.query.get(current_user_id)
    
    assignment = Assignment.query.get_or_404(assignment_id)
    
    if assignment.workspace_id != user.workspace_id:
        return jsonify({"error": "Unauthorized"}), 403
        
    if user.role == 'student' and assignment.status != 'published':
        return jsonify({"error": "Assignment not published"}), 403
        
    return jsonify(assignment.to_dict())

@assignments_bp.route('/<int:assignment_id>', methods=['PUT'])
@jwt_required()
@audit_logger('ASSIGNMENT_UPDATE', 'assignment')
def update_assignment(assignment_id):
    current_user_id = get_jwt_identity()
    user = User.query.get(current_user_id)
    assignment = Assignment.query.get_or_404(assignment_id)
    
    if assignment.workspace_id != user.workspace_id or user.role == 'student':
        return jsonify({"error": "Unauthorized"}), 403
        
    data = request.get_json()
    
    if 'title' in data: assignment.title = data['title']
    if 'description' in data: assignment.description = data['description']
    if 'channel_id' in data: assignment.channel_id = data['channel_id']
    if 'due_date' in data: 
         val = data['due_date']
         if val and isinstance(val, str):
             val = val.replace('Z', '+00:00')
         assignment.due_date = datetime.fromisoformat(val) if val else None
    if 'status' in data: assignment.status = data['status']
    if 'settings' in data: assignment.settings = json.dumps(data['settings'])
    
    db.session.commit()
    return jsonify(assignment.to_dict())

@assignments_bp.route('/<int:assignment_id>/groups', methods=['GET'])
@jwt_required()
def get_all_assignment_groups(assignment_id):
    """List all groups for a specific assignment"""
    current_user_id = get_jwt_identity()
    user = User.query.get(current_user_id)
    assignment = Assignment.query.get_or_404(assignment_id)
    
    if assignment.workspace_id != user.workspace_id:
        return jsonify({"error": "Unauthorized"}), 403
        
    groups = AssignmentGroup.query.filter_by(assignment_id=assignment.id).order_by(AssignmentGroup.id.asc()).all()
    # Also check if the current user is in any of these groups
    user_memberships = AssignmentGroupMember.query.filter_by(user_id=user.id).all()
    user_group_ids = [m.group_id for m in user_memberships]
    
    results = []
    for g in groups:
        g_dict = g.to_dict()
        g_dict['is_member'] = g.id in user_group_ids
        results.append(g_dict)
        
    return jsonify(results)

@assignments_bp.route('/groups/<int:group_id>', methods=['GET'])
@jwt_required()
def get_assignment_group_details(group_id):
    """Get details of a specific assignment group"""
    current_user_id = get_jwt_identity()
    user = User.query.get(current_user_id)
    group = AssignmentGroup.query.get_or_404(group_id)
    
    if group.workspace_id != user.workspace_id:
        return jsonify({"error": "Unauthorized"}), 403
        
    is_member = AssignmentGroupMember.query.filter_by(group_id=group.id, user_id=user.id).first()
    if not is_member and user.role == 'student':
         return jsonify({"error": "Unauthorized"}), 403

    group_dict = group.to_dict()
    group_dict['assignment_title'] = group.assignment.title if group.assignment else "Assignment"
    return jsonify(group_dict)

# -------------------------------------------------------------------
# GROUP MANAGEMENT (GSR)
# -------------------------------------------------------------------

@assignments_bp.route('/<int:assignment_id>/groups', methods=['POST'])
@jwt_required()
def create_groups(assignment_id):
    """
    Create groups for an assignment.
    Body:
    - mode: 'manual' | 'random' | 'count'
    - groups: [{ name: 'Group 1', members: [id1, id2] }] (manual)
    - count: 5 (random/count)
    """
    current_user_id = get_jwt_identity()
    user = User.query.get(current_user_id)
    assignment = Assignment.query.get_or_404(assignment_id)
    
    if assignment.workspace_id != user.workspace_id or user.role == 'student':
        return jsonify({"error": "Unauthorized"}), 403
        
    # Prevent duplicate group creation if they already exist
    existing_groups_count = AssignmentGroup.query.filter_by(assignment_id=assignment.id).count()
    if existing_groups_count > 0:
        return jsonify({"message": "Groups already exist for this assignment", "count": existing_groups_count}), 200

    data = request.get_json()
    mode = data.get('mode', 'manual')
    
    def safe_int(val, default):
        try:
            return int(val) if val is not None else default
        except (ValueError, TypeError):
            return default

    max_members = safe_int(data.get('group_size'), 5)
    if max_members < 1: max_members = 5
    created_groups = []
    
    if mode == 'self_signup':
        count = safe_int(data.get('group_count'), 1)
        for i in range(count):
            group = AssignmentGroup(
                name=f"Group {i+1}",
                assignment_id=assignment.id,
                workspace_id=user.workspace_id,
                max_members=max_members
            )
            db.session.add(group)
            created_groups.append(group)
            
    elif mode == 'manual':
        manual_groups_list = data.get('manual_groups')
        groups_input = data.get('groups', [])
        
        if manual_groups_list:
             for i, member_ids in enumerate(manual_groups_list):
                 if not member_ids: continue
                 group = AssignmentGroup(
                    name=f"Group {i+1}",
                    assignment_id=assignment.id,
                    workspace_id=user.workspace_id
                 )
                 db.session.add(group)
                 db.session.flush()
                 for uid in member_ids:
                     member = AssignmentGroupMember(group_id=group.id, user_id=uid)
                     db.session.add(member)
                 created_groups.append(group)
        else:
            for grp_data in groups_input:
                group = AssignmentGroup(
                    name=grp_data.get('name', 'New Group'),
                    assignment_id=assignment.id,
                    workspace_id=user.workspace_id
                )
                db.session.add(group)
                db.session.flush() # get ID
                
                for uid in grp_data.get('members', []):
                    member = AssignmentGroupMember(group_id=group.id, user_id=uid)
                    db.session.add(member)
                
                created_groups.append(group)
            
    elif mode == 'count':
        count = safe_int(data.get('count'), 1)
        for i in range(count):
            group = AssignmentGroup(
                name=f"Group {i+1}",
                assignment_id=assignment.id,
                workspace_id=user.workspace_id
            )
            db.session.add(group)
            created_groups.append(group)

    elif mode == 'random':
        max_members = safe_int(data.get('group_size'), 5)
        target_student_ids = data.get('student_ids')
        if not target_student_ids:
             if assignment.channel_id:
                 from app.models import ChannelMember
                 members = db.session.query(ChannelMember).join(User).filter(
                     ChannelMember.channel_id == assignment.channel_id,
                     User.role == 'student'
                 ).all()
                 target_student_ids = [m.user_id for m in members]
             else:
                 # Fallback to workspace-wide students if no channel is linked
                 workspace_students = User.query.filter_by(
                     workspace_id=assignment.workspace_id, 
                     role='student'
                 ).all()
                 target_student_ids = [s.id for s in workspace_students]
        
        if not target_student_ids:
             return jsonify({"error": "No students found for grouping. Please select students or link assignment to a channel."}), 400

        shuffled_ids = list(target_student_ids)
        random.shuffle(shuffled_ids)
        
        group_count = data.get('group_count')
        group_size = data.get('group_size')
        
        chunks = []
        if group_count:
             group_count = safe_int(group_count, 1)
             groups_lists = [[] for _ in range(group_count)]
             for i, uid in enumerate(shuffled_ids):
                 groups_lists[i % group_count].append(uid)
             chunks = groups_lists
        elif group_size:
             group_size = safe_int(group_size, 5)
             for i in range(0, len(shuffled_ids), group_size):
                 chunks.append(shuffled_ids[i:i + group_size])
        else:
             chunks = [shuffled_ids]
             
        for i, members in enumerate(chunks):
            if not members: continue
            group = AssignmentGroup(
                name=f"Group {i+1}",
                assignment_id=assignment.id,
                workspace_id=user.workspace_id
            )
            db.session.add(group)
            db.session.flush()
            
            for uid in members:
                mem = AssignmentGroupMember(group_id=group.id, user_id=int(uid))
                db.session.add(mem)
            created_groups.append(group)
            
    # Update max_members for all created groups if specified
    for g in created_groups:
        g.max_members = max_members
            
    db.session.commit()
    return jsonify([g.to_dict() for g in created_groups]), 201

@assignments_bp.route('/groups/<int:group_id>/join', methods=['POST'])
@jwt_required()
def join_assignment_group(group_id):
    """Students join an assignment group manually"""
    current_user_id = get_jwt_identity()
    user = User.query.get(current_user_id)
    group = AssignmentGroup.query.get_or_404(group_id)
    
    # Check if student is in this workspace
    if group.workspace_id != user.workspace_id:
        return jsonify({"error": "Unauthorized"}), 403
        
    # Check if user is already in A group for THIS assignment
    existing_membership = AssignmentGroupMember.query.join(AssignmentGroup).filter(
        AssignmentGroup.assignment_id == group.assignment_id,
        AssignmentGroupMember.user_id == user.id
    ).first()
    
    if existing_membership:
        return jsonify({"error": "You are already a member of a group for this assignment"}), 400
        
    # Check capacity
    if group.members.count() >= group.max_members:
        return jsonify({"error": "This group is full"}), 400
        
    # Check if room is locked
    if group.is_locked:
        return jsonify({"error": "Group joining is locked"}), 403
        
    new_member = AssignmentGroupMember(group_id=group.id, user_id=user.id)
    db.session.add(new_member)
    db.session.commit()
    
    return jsonify({"message": "Successfully joined group", "group": group.to_dict()}), 201

@assignments_bp.route('/<int:assignment_id>/launch-rooms', methods=['POST'])
@jwt_required()
def launch_rooms(assignment_id):
    current_user_id = get_jwt_identity()
    assignment = Assignment.query.get_or_404(assignment_id)
    
    user = User.query.get(current_user_id)
    is_owner = assignment.created_by == int(current_user_id)
    is_staff = user and user.role in ['admin', 'teacher'] and user.workspace_id == assignment.workspace_id
    
    if not (is_owner or is_staff):
         return jsonify({"error": "Unauthorized"}), 403

    asg_groups = AssignmentGroup.query.filter_by(assignment_id=assignment.id).all()
    created_count = 0
    
    for ag in asg_groups:
        room_name = f"{assignment.title} - {ag.name}"
        # Check dupe
        if Group.query.filter_by(name=room_name, workspace_id=assignment.workspace_id).first():
            continue
            
        new_group = Group(
            name=room_name,
            description=f"Study Room for {assignment.title}",
            category='students',
            created_by=current_user_id,
            workspace_id=assignment.workspace_id,
            is_active=True,
            join_type='direct' # Direct so members are added
        )
        db.session.add(new_group)
        db.session.flush()
        
        # Members
        for m in ag.members:
            db.session.add(GroupMember(group_id=new_group.id, user_id=m.user_id, role='member'))
        
        # Owner
        db.session.add(GroupMember(group_id=new_group.id, user_id=current_user_id, role='admin'))
        created_count += 1
        
    db.session.commit()
    return jsonify({"message": "Rooms launched", "count": created_count}), 201

@assignments_bp.route('/<int:assignment_id>/my-group', methods=['GET'])
@jwt_required()
def get_my_group(assignment_id):
    """Get the group the current user belongs to for this assignment"""
    current_user_id = get_jwt_identity()
    
    # optimized query
    membership = AssignmentGroupMember.query.join(AssignmentGroup).filter(
        AssignmentGroup.assignment_id == assignment_id,
        AssignmentGroupMember.user_id == current_user_id
    ).first()
    
    if not membership:
        # Check if teacher (can see all or just return empty?)
        # For now return 404 for student
        return jsonify({"message": "Not assigned to a group"}), 404
        
    group = membership.group
    return jsonify(group.to_dict())

@assignments_bp.route('/my-groups', methods=['GET'])
@jwt_required()
def get_my_groups():
    """List all GSRs the current user belongs to"""
    current_user_id = get_jwt_identity()
    memberships = AssignmentGroupMember.query.filter_by(user_id=current_user_id).all()
    
    results = []
    for m in memberships:
        # Check if group and assignment exist (consistency)
        if m.group and m.group.assignment:
             grp_dict = m.group.to_dict()
             asg_dict = m.group.assignment.to_dict()
             results.append({
                 "group": grp_dict,
                 "assignment": asg_dict,
                 "role": m.role,
                 "joined_at": m.joined_at.isoformat()
             })
    return jsonify(results)

@assignments_bp.route('/groups/<int:group_id>/token', methods=['POST'])
@jwt_required()
def get_group_room_token(group_id):
    """Get LiveKit token for the Group Study Room"""
    current_user_id = get_jwt_identity()
    user = User.query.get(current_user_id)
    group = AssignmentGroup.query.get_or_404(group_id)
    
    # Authorization Check
    is_member = AssignmentGroupMember.query.filter_by(group_id=group.id, user_id=user.id).first()
    is_teacher = (user.role in ['teacher', 'admin', 'super_admin'] and group.workspace_id == user.workspace_id)
    
    if not is_member and not is_teacher:
         return jsonify({"error": "Not a member of this group"}), 403
         
    if group.is_locked and not is_teacher:
         return jsonify({"error": "Group room is locked"}), 403

    # Generate Token
    # Room Name Format: ws_{ws_id}_asg_{asg_id}_grp_{grp_id}
    room_name = f"ws_{group.workspace_id}_asg_{group.assignment_id}_grp_{group.id}"
    
    lk_service = get_livekit_service()
    participant_name = (f"{user.first_name or ''} {user.last_name or ''}".strip()) or user.username
    
    try:
        token = lk_service.create_token(
            room_name=room_name,
            participant_identity=str(user.id),
            participant_name=participant_name,
            is_admin=is_teacher
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500
        
    return jsonify({
        "token": token,
        "url": lk_service.get_public_host(request),
        "room_name": room_name
    })

# -------------------------------------------------------------------
# GSR COLLABORATION (MESSAGES)
# -------------------------------------------------------------------

@assignments_bp.route('/groups/<int:group_id>/messages', methods=['GET'])
@jwt_required()
def get_assignment_group_messages(group_id):
    from app.models import AssignmentGroupMessage
    current_user_id = get_jwt_identity()
    group = AssignmentGroup.query.get_or_404(group_id)
    
    # Simple check: user belongs to the group
    is_member = AssignmentGroupMember.query.filter_by(group_id=group.id, user_id=current_user_id).first()
    if not is_member and User.query.get(current_user_id).role not in ['admin', 'super_admin']:
        return jsonify({"error": "Unauthorized"}), 403
        
    messages = AssignmentGroupMessage.query.filter_by(group_id=group_id).order_by(AssignmentGroupMessage.created_at.asc()).all()
    return jsonify([m.to_dict() for m in messages])

@assignments_bp.route('/groups/<int:group_id>/messages', methods=['POST'])
@jwt_required()
def create_assignment_group_message(group_id):
    from app.models import AssignmentGroupMessage
    current_user_id = get_jwt_identity()
    group = AssignmentGroup.query.get_or_404(group_id)
    
    data = request.get_json()
    if not data.get('content'):
        return jsonify({"error": "Content required"}), 400
        
    msg = AssignmentGroupMessage(
        group_id=group_id,
        user_id=current_user_id,
        content=data.get('content'),
        message_type=data.get('message_type', 'text')
    )
    db.session.add(msg)
    db.session.commit()
    
    return jsonify(msg.to_dict()), 201

# -------------------------------------------------------------------
# GSR COLLABORATION (TASKS)
# -------------------------------------------------------------------

@assignments_bp.route('/groups/<int:group_id>/tasks', methods=['GET'])
@jwt_required()
def get_assignment_group_tasks(group_id):
    from app.models import GroupTask
    current_user_id = get_jwt_identity()
    group = AssignmentGroup.query.get_or_404(group_id)
    
    # Auth check
    is_member = AssignmentGroupMember.query.filter_by(group_id=group.id, user_id=current_user_id).first()
    if not is_member and User.query.get(current_user_id).role not in ['admin', 'super_admin']:
        return jsonify({"error": "Unauthorized"}), 403
        
    tasks = GroupTask.query.filter_by(group_id=group_id).order_by(GroupTask.created_at.desc()).all()
    return jsonify([t.to_dict() for t in tasks])

@assignments_bp.route('/groups/<int:group_id>/tasks', methods=['POST'])
@jwt_required()
def create_assignment_group_task(group_id):
    from app.models import GroupTask
    current_user_id = get_jwt_identity()
    group = AssignmentGroup.query.get_or_404(group_id)
    
    data = request.get_json()
    task = GroupTask(
        group_id=group_id,
        title=data.get('title'),
        description=data.get('description'),
        status=data.get('status', 'To Do'),
        assigned_to=data.get('assigned_to')
    )
    db.session.add(task)
    db.session.commit()
    
    return jsonify(task.to_dict()), 201

@assignments_bp.route('/groups/tasks/<int:task_id>', methods=['PUT'])
@jwt_required()
def update_group_task(task_id):
    from app.models import GroupTask
    task = GroupTask.query.get_or_404(task_id)
    data = request.get_json()
    
    if 'title' in data: task.title = data['title']
    if 'description' in data: task.description = data['description']
    if 'status' in data: task.status = data['status']
    if 'assigned_to' in data: task.assigned_to = data['assigned_to']
    
    db.session.commit()
    return jsonify(task.to_dict())

@assignments_bp.route('/groups/tasks/<int:task_id>', methods=['DELETE'])
@jwt_required()
def delete_group_task(task_id):
    from app.models import GroupTask
    task = GroupTask.query.get_or_404(task_id)
    db.session.delete(task)
    db.session.commit()
    return jsonify({"message": "Task deleted"})

# -------------------------------------------------------------------
# GSR MEMBER MANAGEMENT
# -------------------------------------------------------------------

@assignments_bp.route('/groups/<int:group_id>/members', methods=['GET'])
@jwt_required()
def get_group_members(group_id):
    current_user_id = get_jwt_identity()
    user = User.query.get(current_user_id)
    group = AssignmentGroup.query.get_or_404(group_id)
    
    if group.workspace_id != user.workspace_id:
        return jsonify({"error": "Unauthorized"}), 403
        
    members = AssignmentGroupMember.query.filter_by(group_id=group_id).all()
    results = []
    for m in members:
        u = User.query.get(m.user_id)
        if u:
            results.append({
                "id": u.id,
                "username": u.username,
                "first_name": u.first_name,
                "last_name": u.last_name,
                "role": u.role,
                "joined_at": m.joined_at.isoformat() if hasattr(m, 'joined_at') and m.joined_at else None
            })
    return jsonify(results)

@assignments_bp.route('/groups/<int:group_id>/members', methods=['POST'])
@jwt_required()
def add_group_members(group_id):
    current_user_id = get_jwt_identity()
    user = User.query.get(current_user_id)
    group = AssignmentGroup.query.get_or_404(group_id)
    
    # Check permissions: At least teacher/admin in the same workspace
    if group.workspace_id != user.workspace_id or user.role not in ['teacher', 'admin', 'super_admin']:
        return jsonify({"error": "Unauthorized"}), 403
        
    data = request.get_json()
    user_ids = data.get('user_ids', [])
    if data.get('user_id'):
        user_ids.append(data.get('user_id'))
        
    added_count = 0
    for uid in user_ids:
        # Check if already a member
        existing = AssignmentGroupMember.query.filter_by(group_id=group_id, user_id=uid).first()
        if not existing:
            # Check if user exists and is in the same workspace
            target_user = User.query.get(uid)
            if target_user and target_user.workspace_id == user.workspace_id:
                new_member = AssignmentGroupMember(group_id=group_id, user_id=uid)
                db.session.add(new_member)
                added_count += 1
                
    db.session.commit()
    return jsonify({"message": f"Added {added_count} members", "added_count": added_count}), 201

@assignments_bp.route('/groups/<int:group_id>/members/<int:member_user_id>', methods=['DELETE'])
@jwt_required()
def remove_group_member(group_id, member_user_id):
    current_user_id = get_jwt_identity()
    user = User.query.get(current_user_id)
    group = AssignmentGroup.query.get_or_404(group_id)
    
    if group.workspace_id != user.workspace_id or user.role not in ['teacher', 'admin', 'super_admin']:
        return jsonify({"error": "Unauthorized"}), 403
        
    member = AssignmentGroupMember.query.filter_by(group_id=group_id, user_id=member_user_id).first()
    if not member:
        return jsonify({"error": "Member not found"}), 404
        
    db.session.delete(member)
    db.session.commit()
    return jsonify({"message": "Member removed"})

@assignments_bp.route("/<int:assignment_id>/generate_report", methods=["POST"])
@jwt_required()
def generate_assignment_report(assignment_id):
    current_user_id = get_jwt_identity()
    user = User.query.get(current_user_id)
    assignment = Assignment.query.get_or_404(assignment_id)
    
    if assignment.created_by != int(current_user_id) and user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403

    groups = AssignmentGroup.query.filter_by(assignment_id=assignment.id).order_by(AssignmentGroup.name.asc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Group Roster"

    # Formatting Styles
    header_font = Font(bold=True, size=14)
    sub_header_font = Font(bold=True, size=10)
    summary_font = Font(italic=True)
    table_header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    center_align = Alignment(horizontal="center")

    # Row 1: Title
    ws.merge_cells("A1:D1")
    ws["A1"] = f"Group Listing for {assignment.title}"
    ws["A1"].font = header_font
    ws["A1"].alignment = center_align

    # Row 2: Generated Date
    ws.merge_cells("A2:D2")
    now = datetime.now()
    ws["A2"] = f"Report Generated: {now.strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].alignment = center_align

    # Row 3: Summary Stats
    total_students = AssignmentGroupMember.query.join(AssignmentGroup).filter(AssignmentGroup.assignment_id == assignment.id).count()
    ws.merge_cells("A3:C3")
    ws["A3"] = f"Total Groups: {len(groups)}"
    ws["D3"] = f"Total Students Enrolled: {total_students}"

    # Row 5: Main Header Row
    headers = ["Group Name / ID", "Student ID", "Student Full Name", "Student Email"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = header
        cell.font = sub_header_font
        cell.fill = table_header_fill

    # Data Rows
    current_row = 6
    for group in groups:
        members = group.members.all()
        first_member = True
        for m in members:
            u = m.user
            if not u: continue
            
            if first_member:
                ws.cell(row=current_row, column=1).value = group.name
                first_member = False
            
            ws.cell(row=current_row, column=2).value = f"S{u.id:04d}"
            ws.cell(row=current_row, column=3).value = f"{u.first_name or ''} {u.last_name or ''}".strip() or u.username
            ws.cell(row=current_row, column=4).value = u.email
            current_row += 1
        
        # Blank row separator
        current_row += 1

    # Footer Row
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    ws.cell(row=current_row, column=1).value = f"* End of Report - {total_students} students listed in {len(groups)} groups. *"
    ws.cell(row=current_row, column=1).font = summary_font
    ws.cell(row=current_row, column=1).alignment = center_align

    # Save logic
    # [CourseCode]_[AssignmentName]_Group_List_[YYYY-MM-DD_HHMM].xlsx
    course_code = "COURSE" # Fallback
    if assignment.channel_id:
        from app.models import Channel
        chan = Channel.query.get(assignment.channel_id)
        if chan and chan.course_code:
            course_code = chan.course_code
            
    timestamp = now.strftime("%Y-%m-%d_%H%M")
    safe_title = "".join([c for c in assignment.title if c.isalnum() or c in (" ", "_", "-")]).strip().replace(" ", "_")
    filename = f"{course_code}_{safe_title}_Group_List_{timestamp}.xlsx"
    unique_filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}"
    
    # Path: uploads/vaults/{user_id}/Assignments/{assignment_id}/
    upload_dir = os.path.join(current_app.root_path, "uploads", "vaults", str(user.id), "Assignments", str(assignment.id))
    os.makedirs(upload_dir, exist_ok=True)
    file_path = os.path.join(upload_dir, unique_filename)
    
    wb.save(file_path)
    file_size = os.path.getsize(file_path)

    # Register in File model
    new_file = File(
        filename=unique_filename,
        original_filename=filename,
        file_path=file_path,
        file_size=file_size,
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        uploaded_by=user.id,
        assignment_id=assignment.id,
        workspace_id=user.workspace_id
    )
    db.session.add(new_file)
    db.session.commit()

    return jsonify(new_file.to_dict()), 201

@assignments_bp.route('/groups/<int:group_id>/join', methods=['POST'])
@jwt_required()
def join_group(group_id):
    current_user_id = get_jwt_identity()
    user = User.query.get(current_user_id)
    group = AssignmentGroup.query.get_or_404(group_id)
    
    # Validation: Workspace
    if group.workspace_id != user.workspace_id:
        return jsonify({"error": "Unauthorized"}), 403
        
    # Check if already in a group for this assignment
    existing = AssignmentGroupMember.query.join(AssignmentGroup).filter(
        AssignmentGroup.assignment_id == group.assignment_id,
        AssignmentGroupMember.user_id == user.id
    ).first()
    
    if existing:
        return jsonify({"error": "Already a member of a group in this assignment"}), 400
        
    # Capacity Check
    current_count = AssignmentGroupMember.query.filter_by(group_id=group.id).count()
    if group.max_members and current_count >= group.max_members:
        return jsonify({"error": "Group is full"}), 400
        
    # Join
    member = AssignmentGroupMember(group_id=group.id, user_id=user.id, role='member')
    db.session.add(member)
    db.session.commit()
    
    return jsonify(member.to_dict()), 201

@assignments_bp.route('/<int:assignment_id>/auto-allocate', methods=['POST'])
@jwt_required()
def auto_allocate_students(assignment_id):
    """Magic: Automatically allocate unassigned students to groups with capacity"""
    current_user_id = get_jwt_identity()
    user = User.query.get(current_user_id)
    assignment = Assignment.query.get_or_404(assignment_id)

    if assignment.workspace_id != user.workspace_id or user.role == 'student':
        return jsonify({"error": "Unauthorized"}), 403

    # 1. Get all students (filter by channel if linked, otherwise use workspace)
    if assignment.channel_id:
        from app.models import ChannelMember
        students_query = db.session.query(User).join(ChannelMember).filter(
            ChannelMember.channel_id == assignment.channel_id,
            User.role == 'student'
        )
    else:
        students_query = User.query.filter_by(
            workspace_id=assignment.workspace_id,
            role='student'
        )
    
    students_in_context = students_query.all()

    # 2. Get students already in groups
    assigned_user_ids = db.session.query(AssignmentGroupMember.user_id).join(AssignmentGroup).filter(
        AssignmentGroup.assignment_id == assignment.id
    ).all()
    assigned_user_ids = [uid[0] for uid in assigned_user_ids]

    # 3. Identify unassigned students
    unassigned_students = [s for s in students_in_context if s.id not in assigned_user_ids]
    if not unassigned_students:
        return jsonify({"message": "All students are already assigned", "count": 0}), 200

    # 4. Get groups that are not full
    groups = AssignmentGroup.query.filter_by(assignment_id=assignment.id).all()
    if not groups:
        return jsonify({"error": "No groups exists. Generate group structure first."}), 400

    import random
    random.shuffle(unassigned_students)
    
    allocated_count = 0
    for group in groups:
        current_members = AssignmentGroupMember.query.filter_by(group_id=group.id).count()
        capacity = (group.max_members or 999) - current_members
        
        while capacity > 0 and unassigned_students:
            student = unassigned_students.pop()
            member = AssignmentGroupMember(group_id=group.id, user_id=student.id, role='member')
            db.session.add(member)
            capacity -= 1
            allocated_count += 1

    db.session.commit()
    return jsonify({
        "message": f"Successfully allocated {allocated_count} students",
        "allocated_count": allocated_count,
        "remaining_unassigned": len(unassigned_students)
    }), 200
